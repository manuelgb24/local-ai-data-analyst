"""Dataset validation, loading and profiling for the MVP."""

from __future__ import annotations

from collections.abc import Callable, Iterable, Iterator
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
import re
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from adapters import DuckDBContext, create_duckdb_context
from adapters.duckdb_adapter import quote_identifier
from application.contracts import DatasetColumn, DatasetProfile, ErrorStage, RunError, RunRequest

from .models import PreparedDataset


DuckDBContextFactory = Callable[[], DuckDBContext]

_SUPPORTED_FORMATS = frozenset({"csv", "xlsx", "parquet"})
_XLSX_BATCH_SIZE = 500
_TYPE_VARCHAR = "VARCHAR"
_TYPE_BOOLEAN = "BOOLEAN"
_TYPE_BIGINT = "BIGINT"
_TYPE_DOUBLE = "DOUBLE"
_TYPE_DATE = "DATE"
_TYPE_TIME = "TIME"
_TYPE_TIMESTAMP = "TIMESTAMP"


def detect_dataset_format(path: str | Path) -> str:
    extension = Path(path).suffix.lower().lstrip(".")
    if extension not in _SUPPORTED_FORMATS:
        raise RunError(
            code="dataset_unsupported_format",
            message="Dataset format is not supported by the MVP",
            stage=ErrorStage.DATASET_PREPARATION,
            details={"dataset_path": str(path), "supported_formats": sorted(_SUPPORTED_FORMATS)},
        )
    return extension


def build_table_name(run_id: str) -> str:
    sanitized = re.sub(r"[^a-zA-Z0-9_]", "_", run_id).strip("_")
    if not sanitized:
        raise ValueError("run_id must contain at least one alphanumeric character")
    return f"dataset_{sanitized.lower()}"


class LocalDatasetPreparer:
    """Prepare a local dataset in DuckDB and build its DatasetProfile."""

    def __init__(self, duckdb_factory: DuckDBContextFactory | None = None) -> None:
        self._duckdb_factory = duckdb_factory or create_duckdb_context

    def __call__(self, request: RunRequest, run_id: str) -> PreparedDataset:
        if not isinstance(request, RunRequest):
            raise TypeError("request must be a RunRequest instance")

        source_path = Path(request.dataset_path)
        if not source_path.exists() or not source_path.is_file():
            raise RunError(
                code="dataset_not_found",
                message="Dataset path does not exist or is not a file",
                stage=ErrorStage.DATASET_PREPARATION,
                details={"dataset_path": request.dataset_path},
            )

        dataset_format = detect_dataset_format(source_path)
        if source_path.stat().st_size == 0:
            raise RunError(
                code="dataset_empty",
                message="Dataset file is empty",
                stage=ErrorStage.DATASET_PREPARATION,
                details={"dataset_path": request.dataset_path, "format": dataset_format},
            )

        table_name = build_table_name(run_id)
        duckdb_context = self._duckdb_factory()

        try:
            self._load_into_duckdb(
                duckdb_context=duckdb_context,
                dataset_path=source_path,
                dataset_format=dataset_format,
                table_name=table_name,
            )
            dataset_profile = self._build_profile(
                duckdb_context=duckdb_context,
                source_path=request.dataset_path,
                dataset_format=dataset_format,
                table_name=table_name,
            )
            return PreparedDataset(dataset_profile=dataset_profile, duckdb_context=duckdb_context)
        except RunError:
            duckdb_context.close()
            raise
        except Exception as exc:
            duckdb_context.close()
            raise RunError(
                code="dataset_read_failed",
                message="Dataset could not be loaded into DuckDB",
                stage=ErrorStage.DATASET_PREPARATION,
                details={
                    "dataset_path": request.dataset_path,
                    "format": dataset_format,
                    "error_type": type(exc).__name__,
                    "error_message": str(exc),
                },
            ) from exc

    def _load_into_duckdb(
        self,
        duckdb_context: DuckDBContext,
        dataset_path: Path,
        dataset_format: str,
        table_name: str,
    ) -> None:
        if dataset_format == "csv":
            duckdb_context.create_table_from_select(
                table_name,
                "SELECT * FROM read_csv_auto(?, HEADER=TRUE)",
                [str(dataset_path)],
            )
            return

        if dataset_format == "parquet":
            duckdb_context.create_table_from_select(
                table_name,
                "SELECT * FROM read_parquet(?)",
                [str(dataset_path)],
            )
            return

        self._load_xlsx(duckdb_context=duckdb_context, dataset_path=dataset_path, table_name=table_name)

    def _build_profile(
        self,
        duckdb_context: DuckDBContext,
        source_path: str,
        dataset_format: str,
        table_name: str,
    ) -> DatasetProfile:
        schema_rows = duckdb_context.describe_table(table_name)
        schema = [
            DatasetColumn(name=str(column_name), type=str(column_type))
            for column_name, column_type, *_ in schema_rows
        ]

        if not schema:
            raise RunError(
                code="dataset_empty",
                message="Dataset does not contain any readable columns",
                stage=ErrorStage.DATASET_PREPARATION,
                details={"dataset_path": source_path, "format": dataset_format},
            )

        row_count = int(
            duckdb_context.fetchone(f"SELECT COUNT(*) FROM {quote_identifier(table_name)}")[0]
        )
        if row_count <= 0:
            raise RunError(
                code="dataset_empty",
                message="Dataset does not contain any data rows",
                stage=ErrorStage.DATASET_PREPARATION,
                details={"dataset_path": source_path, "format": dataset_format},
            )

        return DatasetProfile(
            source_path=source_path,
            format=dataset_format,
            table_name=table_name,
            schema=schema,
            row_count=row_count,
        )

    def _load_xlsx(self, duckdb_context: DuckDBContext, dataset_path: Path, table_name: str) -> None:
        try:
            workbook = load_workbook(filename=dataset_path, read_only=True, data_only=True)
        except (InvalidFileException, BadZipFile, OSError, ValueError) as exc:
            raise RunError(
                code="dataset_read_failed",
                message="XLSX dataset could not be opened",
                stage=ErrorStage.DATASET_PREPARATION,
                details={
                    "dataset_path": str(dataset_path),
                    "format": "xlsx",
                    "error_type": type(exc).__name__,
                    "error_message": str(exc),
                },
            ) from exc

        try:
            try:
                worksheet = workbook.worksheets[0]
            except IndexError as exc:
                raise RunError(
                    code="dataset_empty",
                    message="XLSX dataset does not contain any worksheet",
                    stage=ErrorStage.DATASET_PREPARATION,
                    details={"dataset_path": str(dataset_path), "format": "xlsx"},
                ) from exc

            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if header_row is None:
                raise RunError(
                    code="dataset_empty",
                    message="XLSX dataset does not contain a header row",
                    stage=ErrorStage.DATASET_PREPARATION,
                    details={"dataset_path": str(dataset_path), "format": "xlsx"},
                )

            column_names = self._extract_header_names(header_row)
            initial_batch = self._collect_xlsx_batch(rows, column_count=len(column_names), batch_size=_XLSX_BATCH_SIZE)
            if not initial_batch:
                raise RunError(
                    code="dataset_empty",
                    message="XLSX dataset does not contain any data rows",
                    stage=ErrorStage.DATASET_PREPARATION,
                    details={"dataset_path": str(dataset_path), "format": "xlsx"},
                )

            column_types = [
                self._infer_duckdb_type(row[index] for row in initial_batch) for index in range(len(column_names))
            ]
            column_definitions = list(zip(column_names, column_types, strict=True))
            duckdb_context.create_table(table_name, column_definitions)

            column_types = self._insert_xlsx_batch(
                duckdb_context=duckdb_context,
                table_name=table_name,
                column_names=column_names,
                column_types=column_types,
                batch_rows=initial_batch,
            )

            while True:
                batch_rows = self._collect_xlsx_batch(rows, column_count=len(column_names), batch_size=_XLSX_BATCH_SIZE)
                if not batch_rows:
                    break
                column_types = self._insert_xlsx_batch(
                    duckdb_context=duckdb_context,
                    table_name=table_name,
                    column_names=column_names,
                    column_types=column_types,
                    batch_rows=batch_rows,
                )
        finally:
            workbook.close()

    def _extract_header_names(self, header_row: tuple[object, ...]) -> list[str]:
        header_values = list(header_row)
        while header_values and self._is_blank(header_values[-1]):
            header_values.pop()

        if not header_values:
            raise RunError(
                code="dataset_empty",
                message="XLSX dataset does not contain any columns",
                stage=ErrorStage.DATASET_PREPARATION,
                details={"format": "xlsx"},
            )

        column_names: list[str] = []
        for index, value in enumerate(header_values, start=1):
            if self._is_blank(value):
                raise RunError(
                    code="dataset_invalid_header",
                    message="XLSX header row must contain column names",
                    stage=ErrorStage.DATASET_PREPARATION,
                    details={"format": "xlsx", "column_index": index},
                )
            column_names.append(str(value).strip())

        return column_names

    def _collect_xlsx_batch(
        self,
        rows: Iterator[tuple[object, ...]],
        column_count: int,
        batch_size: int,
    ) -> list[tuple[object, ...]]:
        prepared_rows: list[tuple[object, ...]] = []
        for row in rows:
            values = list(row[:column_count])
            if len(values) < column_count:
                values.extend([None] * (column_count - len(values)))

            if all(self._is_blank(value) for value in values):
                continue

            prepared_rows.append(tuple(self._normalize_cell(value) for value in values))

            if len(prepared_rows) >= batch_size:
                break

        return prepared_rows

    def _insert_xlsx_batch(
        self,
        duckdb_context: DuckDBContext,
        table_name: str,
        column_names: list[str],
        column_types: list[str],
        batch_rows: list[tuple[object, ...]],
    ) -> list[str]:
        promoted_types = list(column_types)
        for row in batch_rows:
            promoted_types = [
                self._promote_duckdb_type(current_type, row[index])
                for index, current_type in enumerate(promoted_types)
            ]

        for index, (current_type, promoted_type) in enumerate(zip(column_types, promoted_types, strict=True)):
            if current_type != promoted_type:
                duckdb_context.alter_column_type(table_name, column_names[index], promoted_type)

        duckdb_context.insert_many(table_name, batch_rows, column_count=len(column_names))
        return promoted_types

    def _infer_duckdb_type(self, values: Iterable[object]) -> str:
        observed = [value for value in values if value is not None]
        if not observed:
            return _TYPE_VARCHAR

        inferred_type = self._value_duckdb_type(observed[0])
        for value in observed[1:]:
            inferred_type = self._merge_duckdb_types(inferred_type, self._value_duckdb_type(value))
        return inferred_type

    def _promote_duckdb_type(self, current_type: str, value: object) -> str:
        if value is None:
            return current_type
        value_type = self._value_duckdb_type(value)
        return self._merge_duckdb_types(current_type, value_type)

    def _value_duckdb_type(self, value: object) -> str:
        if value is None:
            return _TYPE_VARCHAR
        if isinstance(value, bool):
            return _TYPE_BOOLEAN
        if isinstance(value, int):
            return _TYPE_BIGINT
        if isinstance(value, (float, Decimal)):
            return _TYPE_DOUBLE
        if isinstance(value, datetime):
            return _TYPE_TIMESTAMP
        if isinstance(value, date):
            return _TYPE_DATE
        if isinstance(value, time):
            return _TYPE_TIME
        return _TYPE_VARCHAR

    def _merge_duckdb_types(self, left: str, right: str) -> str:
        if left == right:
            return left
        if _TYPE_VARCHAR in {left, right}:
            return _TYPE_VARCHAR

        numeric_types = {left, right}
        if numeric_types <= {_TYPE_BOOLEAN, _TYPE_BIGINT}:
            return _TYPE_BIGINT
        if numeric_types <= {_TYPE_BOOLEAN, _TYPE_BIGINT, _TYPE_DOUBLE}:
            return _TYPE_DOUBLE
        if {left, right} <= {_TYPE_DATE, _TYPE_TIME, _TYPE_TIMESTAMP}:
            return _TYPE_VARCHAR
        return _TYPE_VARCHAR

    def _normalize_cell(self, value: object) -> object:
        if self._is_blank(value):
            return None
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _is_blank(self, value: object) -> bool:
        return value is None or (isinstance(value, str) and not value.strip())
