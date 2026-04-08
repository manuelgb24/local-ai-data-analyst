from pathlib import Path
from datetime import date

from openpyxl import Workbook
import pytest

from application import AgentExecutionContext, AgentResult, ArtifactManifest, RunError, RunRequest, RunState
from data import LocalDatasetPreparer
from runtime import AgentRegistry, InMemoryRunTracker, RegisteredAgent, RuntimeCoordinator


def write_csv(path: Path) -> Path:
    path.write_text("store,sales,active\n1,100.5,true\n2,200.0,false\n", encoding="utf-8")
    return path


def write_xlsx(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["store", "sales", "day"])
    worksheet.append([1, 100.5, "2024-01-01"])
    worksheet.append([2, 200.0, "2024-01-02"])
    workbook.save(path)
    return path


def write_xlsx_mixed_types(path: Path, row_count: int = 3) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "primary"
    worksheet.append(["store", "mixed_value", "event_date"])
    worksheet.append([1, 100, date(2024, 1, 1)])
    worksheet.append([2, "fallback-to-text", date(2024, 1, 2)])
    for index in range(3, row_count + 1):
        worksheet.append([index, index * 10, date(2024, 1, min(index, 28))])

    second = workbook.create_sheet("secondary")
    second.append(["should", "be", "ignored"])
    second.append(["x", "y", "z"])
    workbook.save(path)
    return path


def write_xlsx_with_nulls_after_first_batch(path: Path, row_count: int = 700) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "primary"
    worksheet.append(["store", "quantity"])
    for index in range(1, row_count + 1):
        quantity = index if index <= 500 else None
        worksheet.append([index, quantity])
    workbook.save(path)
    return path


def write_parquet(path: Path) -> Path:
    csv_path = write_csv(path.with_suffix(".csv"))
    preparer = LocalDatasetPreparer()
    prepared = preparer(
        RunRequest(
            agent_id="data_analyst",
            dataset_path=str(csv_path),
            user_prompt="Preparar parquet",
        ),
        run_id="run-source-for-parquet",
    )
    prepared.duckdb_context.execute(
        f"COPY {prepared.dataset_profile.table_name} TO ? (FORMAT PARQUET)",
        [str(path)],
    )
    prepared.duckdb_context.close()
    return path


def build_request(dataset_path: Path) -> RunRequest:
    return RunRequest(
        agent_id="data_analyst",
        dataset_path=str(dataset_path),
        user_prompt="Resume las ventas",
    )


def build_registry(agent_executor) -> AgentRegistry:
    return AgentRegistry(
        {
            "data_analyst": RegisteredAgent(
                agent_id="data_analyst",
                executor=agent_executor,
                config={"model": "deepseek-r1:8b"},
            )
        }
    )


@pytest.mark.parametrize(
    ("writer", "expected_format"),
    [
        (write_csv, "csv"),
        (write_xlsx, "xlsx"),
        (write_parquet, "parquet"),
    ],
    ids=["csv", "xlsx", "parquet"],
)
def test_local_dataset_preparer_loads_supported_formats(
    repo_tmp_path: Path,
    writer,
    expected_format: str,
) -> None:
    dataset_path = writer(repo_tmp_path / f"sample.{expected_format}")
    preparer = LocalDatasetPreparer()

    prepared = preparer(build_request(dataset_path), run_id=f"run-{expected_format}")

    assert prepared.dataset_profile.source_path == str(dataset_path)
    assert prepared.dataset_profile.format == expected_format
    assert prepared.dataset_profile.table_name == f"dataset_run_{expected_format}"
    assert prepared.dataset_profile.row_count == 2
    assert [column.name for column in prepared.dataset_profile.schema][:2] == ["store", "sales"]
    assert prepared.duckdb_context.fetchone(
        f"SELECT COUNT(*) FROM {prepared.dataset_profile.table_name}"
    ) == (2,)

    prepared.duckdb_context.close()


def test_local_dataset_preparer_rejects_corrupt_xlsx(repo_tmp_path: Path) -> None:
    dataset_path = repo_tmp_path / "corrupt.xlsx"
    dataset_path.write_text("not-a-real-xlsx", encoding="utf-8")
    preparer = LocalDatasetPreparer()

    with pytest.raises(RunError, match="dataset_read_failed") as exc_info:
        preparer(build_request(dataset_path), run_id="run-corrupt-xlsx")

    assert exc_info.value.stage.value == "dataset_preparation"


def test_local_dataset_preparer_rejects_corrupt_parquet(repo_tmp_path: Path) -> None:
    dataset_path = repo_tmp_path / "corrupt.parquet"
    dataset_path.write_text("not-a-real-parquet", encoding="utf-8")
    preparer = LocalDatasetPreparer()

    with pytest.raises(RunError, match="dataset_read_failed") as exc_info:
        preparer(build_request(dataset_path), run_id="run-corrupt-parquet")

    assert exc_info.value.stage.value == "dataset_preparation"


def test_local_dataset_preparer_streams_xlsx_batches_and_promotes_mixed_column_to_varchar(repo_tmp_path: Path) -> None:
    dataset_path = write_xlsx_mixed_types(repo_tmp_path / "mixed.xlsx", row_count=1100)
    preparer = LocalDatasetPreparer()

    prepared = preparer(build_request(dataset_path), run_id="run-mixed-xlsx")

    assert prepared.dataset_profile.format == "xlsx"
    assert prepared.dataset_profile.row_count == 1100
    schema_by_name = {column.name: column.type for column in prepared.dataset_profile.schema}
    assert schema_by_name["store"] == "BIGINT"
    assert schema_by_name["mixed_value"] == "VARCHAR"
    assert schema_by_name["event_date"] in {"DATE", "TIMESTAMP"}
    assert prepared.duckdb_context.fetchone(
        f"SELECT mixed_value FROM {prepared.dataset_profile.table_name} WHERE store = 2"
    ) == ("fallback-to-text",)
    prepared.duckdb_context.close()


def test_local_dataset_preparer_uses_only_first_xlsx_sheet(repo_tmp_path: Path) -> None:
    dataset_path = write_xlsx_mixed_types(repo_tmp_path / "first-sheet-only.xlsx", row_count=3)
    preparer = LocalDatasetPreparer()

    prepared = preparer(build_request(dataset_path), run_id="run-first-sheet")

    assert prepared.dataset_profile.row_count == 3
    assert prepared.duckdb_context.fetchone(
        f"SELECT COUNT(*) FROM {prepared.dataset_profile.table_name} WHERE mixed_value = 'x'"
    ) == (0,)
    prepared.duckdb_context.close()


def test_local_dataset_preparer_keeps_numeric_type_when_later_xlsx_batches_are_null(repo_tmp_path: Path) -> None:
    dataset_path = write_xlsx_with_nulls_after_first_batch(repo_tmp_path / "nulls-later.xlsx")
    preparer = LocalDatasetPreparer()

    prepared = preparer(build_request(dataset_path), run_id="run-nulls-later")

    schema_by_name = {column.name: column.type for column in prepared.dataset_profile.schema}
    assert schema_by_name["quantity"] == "BIGINT"
    assert prepared.duckdb_context.fetchone(
        f"SELECT COUNT(*) FROM {prepared.dataset_profile.table_name} WHERE quantity IS NULL"
    ) == (200,)
    prepared.duckdb_context.close()


def test_runtime_coordinator_uses_real_dataset_preparer(repo_tmp_path: Path) -> None:
    tracker = InMemoryRunTracker()
    dataset_path = write_csv(repo_tmp_path / "runtime.csv")
    observed: dict[str, AgentExecutionContext] = {}

    def fake_agent_executor(request: RunRequest, context: AgentExecutionContext) -> AgentResult:
        observed["context"] = context
        return AgentResult(
            narrative="ok",
            findings=[],
            sql_trace=[],
            tables=[],
            charts=[],
            artifact_manifest=ArtifactManifest(run_id=context.run_id),
        )

    coordinator = RuntimeCoordinator(
        dataset_preparer=LocalDatasetPreparer(),
        agent_registry=build_registry(fake_agent_executor),
        tracker=tracker,
        artifacts_root=repo_tmp_path / "artifacts",
    )

    result = coordinator.run(build_request(dataset_path))
    record = tracker.get(result.artifact_manifest.run_id)
    context = observed["context"]

    assert record.state is RunState.SUCCEEDED
    assert context.dataset_profile.row_count == 2
    with pytest.raises(Exception, match="closed"):
        context.duckdb_context.fetchone(f"SELECT SUM(sales) FROM {context.dataset_profile.table_name}")


def test_runtime_coordinator_marks_failed_with_real_dataset_preparer_on_missing_file(repo_tmp_path: Path) -> None:
    tracker = InMemoryRunTracker()
    captured_context_calls: list[AgentExecutionContext] = []

    def fake_agent_executor(request: RunRequest, context: AgentExecutionContext) -> AgentResult:
        captured_context_calls.append(context)
        raise AssertionError("agent executor should not be called")

    coordinator = RuntimeCoordinator(
        dataset_preparer=LocalDatasetPreparer(),
        agent_registry=build_registry(fake_agent_executor),
        tracker=tracker,
        artifacts_root=repo_tmp_path / "artifacts",
    )
    request = build_request(repo_tmp_path / "missing.csv")

    with pytest.raises(RunError, match="dataset_not_found"):
        coordinator.run(request)

    assert captured_context_calls == []
    run_id = next(iter(path.name for path in (repo_tmp_path / "artifacts").iterdir() if path.is_dir()))
    assert tracker.get(run_id).state is RunState.FAILED


def test_runtime_coordinator_marks_failed_with_real_dataset_preparer_on_empty_file(repo_tmp_path: Path) -> None:
    tracker = InMemoryRunTracker()
    captured_context_calls: list[AgentExecutionContext] = []
    dataset_path = repo_tmp_path / "empty.csv"
    dataset_path.write_text("", encoding="utf-8")

    def fake_agent_executor(request: RunRequest, context: AgentExecutionContext) -> AgentResult:
        captured_context_calls.append(context)
        raise AssertionError("agent executor should not be called")

    coordinator = RuntimeCoordinator(
        dataset_preparer=LocalDatasetPreparer(),
        agent_registry=build_registry(fake_agent_executor),
        tracker=tracker,
        artifacts_root=repo_tmp_path / "artifacts",
    )
    request = build_request(dataset_path)

    with pytest.raises(RunError, match="dataset_empty") as exc_info:
        coordinator.run(request)

    assert captured_context_calls == []
    run_id = next(iter(path.name for path in (repo_tmp_path / "artifacts").iterdir() if path.is_dir()))
    record = tracker.get(run_id)
    assert record.state is RunState.FAILED
    assert record.error == exc_info.value
    assert exc_info.value.stage.value == "dataset_preparation"


def test_local_dataset_preparer_profiles_datasetv1_reference_csv() -> None:
    preparer = LocalDatasetPreparer()
    dataset_path = Path("DatasetV1/Walmart_Sales.csv")

    prepared = preparer(build_request(dataset_path), run_id="run-datasetv1")

    assert prepared.dataset_profile.source_path == str(dataset_path)
    assert prepared.dataset_profile.format == "csv"
    assert prepared.dataset_profile.row_count > 0
    assert len(prepared.dataset_profile.schema) > 0
    prepared.duckdb_context.close()
